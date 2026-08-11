"""
clean_stimulation_logs.py

Batch-cleans raw stimulation-log Excel files.

What it does
------------
Each raw file has rows like:

    20:40:55  Start Stimulation from PVNH'1 to PVNH'3 ( F=1 Hz, PD=500 µs,
              TD=30.0 s, I=2.0 mA)

mixed in with junk rows ("De-block start", "01-Mar", "2,MS", etc).

For every row that matches a "Start Stimulation ..." line, this script pulls out:
    - the timestamp (from column A)
    - the first electrode name  (e.g. PVNH'1)
    - the second electrode name (e.g. PVNH'3)
    - the current value + unit  (e.g. "2 mA")

and writes ONLY those rows, in order, to a new clean file in the output folder.
Every non-matching row is dropped.

Usage
-----
    python clean_stimulation_logs.py --input_dir "path/to/raw_files" --output_dir "path/to/clean_files"

If you don't pass --input_dir / --output_dir, it defaults to "./input" and "./output"
(created next to this script).

Requirements
------------
    pip install openpyxl
"""

import argparse
import re
from pathlib import Path
from datetime import datetime, time, date

from openpyxl import load_workbook, Workbook

# Matches: Start Stimulation from <elec1> to <elec2> ( ... I=<number> <unit> )
STIM_PATTERN = re.compile(
    r"Start\s+Stimulation\s+from\s+(\S+)\s+to\s+(\S+).*?"
    r"I\s*=\s*([\d.]+)\s*(m?A|µA|uA)",
    re.IGNORECASE,
)


def format_timestamp(value):
    """Return a clean string for whatever Excel put in column A."""
    if value is None:
        return ""
    if isinstance(value, (datetime, time)):
        return value.strftime("%H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def format_current(number_str, unit):
    """Turn '2.0' -> '2', '6.5' -> '6.5', and pair it with the unit."""
    num = float(number_str)
    if num.is_integer():
        num_str = str(int(num))
    else:
        num_str = str(num)
    return f"{num_str} {unit}"


def extract_rows(sheet):
    """Scan one worksheet and return a list of [timestamp, elec1, elec2, current] rows."""
    cleaned = []
    for row in sheet.iter_rows():
        if not row:
            continue
        timestamp_cell = row[0]
        # Join the text of every other cell in the row so the match works
        # no matter which column the log text landed in.
        row_text = " ".join(
            str(cell.value) for cell in row[1:] if cell.value is not None
        )
        if not row_text:
            continue

        match = STIM_PATTERN.search(row_text)
        if not match:
            continue  # junk row (De-block start, stray notes, etc.) -> skip

        elec1, elec2, current_num, unit = match.groups()
        cleaned.append(
            [
                format_timestamp(timestamp_cell.value),
                elec1,
                elec2,
                format_current(current_num, unit),
            ]
        )
    return cleaned


def clean_file(input_path: Path, output_path: Path):
    wb = load_workbook(input_path, data_only=True)
    sheet = wb.active  # only the active/first sheet is processed

    cleaned_rows = extract_rows(sheet)

    out_wb = Workbook()
    out_sheet = out_wb.active
    out_sheet.title = "Cleaned"
    out_sheet.append(["Timestamp", "Electrode 1", "Electrode 2", "Current"])
    for r in cleaned_rows:
        out_sheet.append(r)

    out_wb.save(output_path)
    return len(cleaned_rows)


def main():
    parser = argparse.ArgumentParser(description="Clean a folder of stimulation-log Excel files.")
    parser.add_argument("--input_dir", default="input", help="Folder containing the raw .xlsx files")
    parser.add_argument("--output_dir", default="output", help="Folder to write cleaned .xlsx files to")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    excel_files = sorted(input_dir.glob("*.xlsx")) + sorted(input_dir.glob("*.xls"))
    if not excel_files:
        print(f"No .xlsx/.xls files found in {input_dir.resolve()}")
        return

    for file_path in excel_files:
        out_path = output_dir / f"{file_path.stem}_clean.xlsx"
        try:
            n = clean_file(file_path, out_path)
            print(f"[OK]   {file_path.name} -> {out_path.name}  ({n} stimulation rows kept)")
        except Exception as e:
            print(f"[FAIL] {file_path.name}: {e}")


if __name__ == "__main__":
    main()
