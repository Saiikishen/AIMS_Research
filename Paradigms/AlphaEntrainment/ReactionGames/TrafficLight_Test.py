import os
import sys
import random
import time
from datetime import datetime

import tkinter as tk
from tkinter import simpledialog, messagebox

import openpyxl
from openpyxl.styles import Font, Alignment

# ---------------------------------------------------------------------------
# Configurable test parameters
# ---------------------------------------------------------------------------
NUM_CYCLES = 5          # number of test cycles
CYCLE_DURATION = 12.0   # seconds, fixed length of each cycle
MIN_GREEN_DELAY = 2.0   # earliest the light can turn green within a cycle
MAX_GREEN_BUFFER = 3.0  # latest = CYCLE_DURATION - this, leaves reaction room
INTER_CYCLE_GAP = 1.0   # short "get ready" pause shown between cycles
EXCEL_FILENAME = "reaction_test_log.xlsx"


class ReactionTimeTest:
    def __init__(self, root):
        self.root = root
        self.root.title("Traffic Light Reaction Time Test")
        self.root.geometry("380x520")
        self.root.resizable(False, False)
        self.root.withdraw()  # hide main window until patient ID is entered

        self.patient_id = self.ask_patient_id()
        if not self.patient_id:
            # user cancelled / gave no ID -> exit cleanly
            self.root.destroy()
            sys.exit(0)

        self.root.deiconify()

        # Test state
        self.cycle_num = 0
        self.results = []          # list of dicts, one per cycle
        self.cycle_state = "idle"  # "waiting_green" -> "green_on" -> handled
        self.responded = False
        self.green_actual_time = None
        self.excel_path = None

        self.build_ui()

        # Capture spacebar presses anywhere in the window
        self.root.bind("<KeyPress-space>", self.on_space_press)
        self.root.focus_force()

        # Small delay before the first cycle so the patient can settle
        self.root.after(1000, self.start_cycle)

    # ------------------------------------------------------------------
    # Setup
    # ------------------------------------------------------------------
    def ask_patient_id(self):
        return simpledialog.askstring(
            "Patient ID",
            "Enter Patient ID:",
            parent=self.root,
        )

    def build_ui(self):
        tk.Label(
            self.root, text=f"Patient ID: {self.patient_id}", font=("Arial", 12)
        ).pack(pady=10)

        self.status_label = tk.Label(
            self.root, text="Get ready...", font=("Arial", 14, "bold")
        )
        self.status_label.pack(pady=8)

        self.canvas = tk.Canvas(self.root, width=160, height=340, bg="black")
        self.canvas.pack(pady=10)

        self.red_light = self.canvas.create_oval(
            30, 15, 130, 115, fill="darkred", outline="white", width=2
        )
        self.yellow_light = self.canvas.create_oval(
            30, 128, 130, 228, fill="#4d4d00", outline="white", width=2
        )
        self.green_light = self.canvas.create_oval(
            30, 241, 130, 341, fill="darkgreen", outline="white", width=2
        )

        self.cycle_label = tk.Label(
            self.root, text=f"Cycle: 0/{NUM_CYCLES}", font=("Arial", 10)
        )
        self.cycle_label.pack(pady=6)

        tk.Label(
            self.root,
            text="Press SPACEBAR the instant the light turns GREEN",
            font=("Arial", 10),
            fg="gray30",
        ).pack(pady=4)

    def set_light(self, color):
        self.canvas.itemconfig(
            self.red_light, fill="red" if color == "red" else "darkred"
        )
        self.canvas.itemconfig(
            self.yellow_light, fill="yellow" if color == "yellow" else "#4d4d00"
        )
        self.canvas.itemconfig(
            self.green_light, fill="#00ff00" if color == "green" else "darkgreen"
        )

    # ------------------------------------------------------------------
    # Cycle logic
    # ------------------------------------------------------------------
    def start_cycle(self):
        if self.cycle_num >= NUM_CYCLES:
            self.finish_test()
            return

        self.cycle_num += 1
        self.cycle_label.config(text=f"Cycle: {self.cycle_num}/{NUM_CYCLES}")
        self.status_label.config(text="Wait for GREEN...", fg="black")
        self.set_light("red")

        self.cycle_state = "waiting_green"
        self.responded = False
        self.green_actual_time = None

        # Random moment (within the 12s window) the light turns green
        green_delay = random.uniform(
            MIN_GREEN_DELAY, CYCLE_DURATION - MAX_GREEN_BUFFER
        )
        self.root.after(int(green_delay * 1000), self.turn_green)

        # Cycle always ends at a fixed 12s mark, regardless of response
        self.root.after(int(CYCLE_DURATION * 1000), self.end_cycle)

    def turn_green(self):
        if self.cycle_state != "waiting_green":
            return  # cycle already resolved (e.g. false start already logged)
        self.cycle_state = "green_on"
        self.set_light("green")
        self.status_label.config(text="PRESS SPACE NOW!", fg="green")
        self.green_actual_time = time.time()

    def on_space_press(self, event):
        if self.cycle_state == "idle" or self.responded:
            return  # ignore presses outside an active cycle / duplicate presses

        if self.cycle_state == "green_on":
            reaction_time = time.time() - self.green_actual_time
            self.responded = True
            self.results.append(
                {
                    "cycle": self.cycle_num,
                    "reaction_time": round(reaction_time, 4),
                    "status": "OK",
                }
            )
            self.status_label.config(
                text=f"Reaction: {reaction_time:.3f}s", fg="blue"
            )
            self.set_light("red")

        elif self.cycle_state == "waiting_green":
            # Pressed before the light turned green
            self.responded = True
            self.results.append(
                {"cycle": self.cycle_num, "reaction_time": None, "status": "FALSE START"}
            )
            self.status_label.config(text="Too early! Wait for green.", fg="red")

    def end_cycle(self):
        if not self.responded:
            self.results.append(
                {"cycle": self.cycle_num, "reaction_time": None, "status": "MISSED"}
            )
            self.status_label.config(text="No response - missed.", fg="orange")

        self.cycle_state = "idle"
        self.set_light("red")
        self.root.after(int(INTER_CYCLE_GAP * 1000), self.start_cycle)

    # ------------------------------------------------------------------
    # Wrap-up
    # ------------------------------------------------------------------
    def finish_test(self):
        self.status_label.config(text="Test Complete!", fg="black")

        valid_times = [
            r["reaction_time"] for r in self.results if r["status"] == "OK"
        ]
        avg = sum(valid_times) / len(valid_times) if valid_times else None

        self.save_to_excel(avg)

        summary_lines = [
            f"Patient ID: {self.patient_id}",
            f"Valid trials: {len(valid_times)}/{NUM_CYCLES}",
        ]
        if avg is not None:
            summary_lines.append(f"Average reaction time: {avg:.3f} s")
        else:
            summary_lines.append("No valid reaction recorded.")
        summary_lines.append(f"\nSaved to:\n{self.excel_path}")

        messagebox.showinfo("Test Complete", "\n".join(summary_lines))
        self.root.after(200, self.root.destroy)

    def save_to_excel(self, avg):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        filepath = os.path.join(script_dir, EXCEL_FILENAME)
        self.excel_path = filepath

        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reaction Test Results"
            headers = [
                "Patient ID",
                "Test Timestamp",
                "Cycle",
                "Reaction Time (s)",
                "Status",
            ]
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            for col_letter in ["A", "B", "C", "D", "E"]:
                ws.column_dimensions[col_letter].width = 20

        test_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for r in self.results:
            ws.append(
                [
                    self.patient_id,
                    test_timestamp,
                    r["cycle"],
                    r["reaction_time"] if r["reaction_time"] is not None else "N/A",
                    r["status"],
                ]
            )

        n_valid = len([r for r in self.results if r["status"] == "OK"])
        ws.append(
            [
                self.patient_id,
                test_timestamp,
                "AVERAGE",
                round(avg, 4) if avg is not None else "N/A",
                f"{n_valid} valid trial(s)",
            ]
        )
        ws.append([])  # blank separator row before the next patient's block

        wb.save(filepath)


if __name__ == "__main__":
    root = tk.Tk()
    app = ReactionTimeTest(root)
    root.mainloop()